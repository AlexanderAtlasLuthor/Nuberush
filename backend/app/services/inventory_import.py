"""Excel (QuickBooks POS) inventory import — F2.27.8.

Three concerns live here, layered so each can be tested in isolation:

  1. Upload plumbing (SUBFASE A): extension / content-type / size
     validation for the raw multipart upload, plus a controlled
     `InventoryImportValidationError` that the router translates into a
     clean HTTP response. None of this touches the DB.

  2. Pure parser (SUBFASE B): `parse_quickbooks_inventory_workbook`
     reads an `.xlsx` exported from QuickBooks POS into a list of
     `ParsedInventoryImportRow`. It is DB-free and side-effect-free —
     it only knows how to turn bytes into normalized rows and to flag
     per-row parse problems (empty SKU, bad quantity). File-level
     problems (unreadable file, no header row, missing required
     columns) raise `InventoryImportValidationError`.

  3. Preview + confirm (SUBFASE D/E): `build_inventory_import_preview`
     diffs parsed rows against existing `ProductVariant` /
     `InventoryItem` rows WITHOUT writing; `confirm_inventory_import`
     re-validates and applies the diff inside a single all-or-nothing
     transaction, reusing the audit + commit helpers from
     `app.services.inventory`.

Scope guarantees (F2.27.8 contract):
  - never creates Product or ProductVariant;
  - only updates/creates `InventoryItem` for variants that ALREADY
    exist (matched by the global-unique `ProductVariant.sku`);
  - preview performs zero DB writes;
  - confirm is transactional: any failing row rolls the whole import
    back;
  - inventory changes are audited via `InventoryLog`
    (movement_type=adjustment, reason="inventory_import"); a row whose
    delta is zero writes no log (the DB CHECK forbids quantity_delta=0).
"""

from __future__ import annotations

import io
from dataclasses import dataclass
from dataclasses import field
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.db.models import InventoryItem
from app.db.models import InventoryMovementType
from app.db.models import InventoryStatus
from app.db.models import ProductVariant
from app.schemas.inventory_import import InventoryImportConfirmResponse
from app.schemas.inventory_import import InventoryImportPreviewResponse
from app.schemas.inventory_import import InventoryImportPreviewRow
from app.schemas.inventory_import import InventoryImportRowIssue
from app.schemas.inventory_import import InventoryImportSummary
from app.services import inventory as inventory_svc


# --------------------------------------------------------------------- #
# Upload plumbing constants (SUBFASE A)
# --------------------------------------------------------------------- #

ALLOWED_INVENTORY_IMPORT_EXTENSIONS: frozenset[str] = frozenset({".xlsx"})

# 10 MB. A QuickBooks POS item export is a few hundred KB even for a
# large catalog; 10 MB leaves generous headroom while still rejecting an
# accidental upload of something enormous.
MAX_INVENTORY_IMPORT_BYTES: int = 10 * 1024 * 1024

# MIME types browsers / OSes attach to a `.xlsx`. The extension is the
# authority (see `validate_inventory_import_content_type`); these only
# gate the cases where a *specific, wrong* type was sent.
INVENTORY_IMPORT_ALLOWED_CONTENT_TYPES: frozenset[str] = frozenset(
    {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",  # some clients send this for .xlsx
        "application/octet-stream",  # generic binary
        "application/zip",  # .xlsx is a zip container
        "",  # empty string from some test clients
    }
)

# Sanity ceiling on a single row's quantity. QuickBooks stores small
# integer counts; anything past this is almost certainly a parse error
# (a price/UPC leaking into the qty column) and is surfaced rather than
# silently written.
MAX_INVENTORY_IMPORT_QUANTITY: int = 1_000_000

# Audit metadata for every inventory mutation produced by an import.
INVENTORY_IMPORT_REASON: str = "inventory_import"


# Row-issue codes (mirrored 1:1 in the frontend types).
CODE_MISSING_SKU = "MISSING_SKU"
CODE_INVALID_QUANTITY = "INVALID_QUANTITY"
CODE_QUANTITY_OUT_OF_RANGE = "QUANTITY_OUT_OF_RANGE"
CODE_DUPLICATE_SKU_IN_FILE = "DUPLICATE_SKU_IN_FILE"
CODE_VARIANT_NOT_FOUND = "VARIANT_NOT_FOUND"
CODE_RESERVED_EXCEEDS_NEW_ONHAND = "RESERVED_EXCEEDS_NEW_ONHAND"

# File-level codes (raised, never per-row).
CODE_BAD_HEADERS = "BAD_HEADERS"
CODE_UNSUPPORTED_FILE_TYPE = "UNSUPPORTED_FILE_TYPE"
CODE_EMPTY_FILE = "EMPTY_FILE"
CODE_FILE_TOO_LARGE = "FILE_TOO_LARGE"
CODE_BLOCKING_ERRORS = "BLOCKING_ERRORS"

# Preview row actions.
ACTION_UPDATE = "update"
ACTION_CREATE_INVENTORY_ITEM = "create_inventory_item"
ACTION_SKIP = "skip"


class InventoryImportValidationError(Exception):
    """A controlled, file-level import failure.

    Carries a stable machine `code` (one of the file-level codes above)
    and an HTTP `status_code` so the router can translate it into a
    clean response without leaking parser internals. Per-row problems
    are NOT raised — they are attached to the preview rows instead.
    """

    def __init__(
        self,
        code: str,
        message: str,
        *,
        status_code: int = 422,
    ) -> None:
        super().__init__(message)
        self.code = code
        self.message = message
        self.status_code = status_code


# --------------------------------------------------------------------- #
# Upload plumbing helpers (SUBFASE A)
# --------------------------------------------------------------------- #


def get_inventory_import_file_extension(filename: str) -> str:
    """Return the lowercased extension (including the dot) of `filename`.

    Returns "" when there is no extension. Case-insensitive so `.XLSX`
    and `.xlsx` are treated identically.
    """
    if not filename:
        return ""
    name = filename.strip()
    dot = name.rfind(".")
    if dot == -1:
        return ""
    return name[dot:].lower()


def validate_inventory_import_filename(filename: str) -> None:
    """Reject anything that is not a non-empty `.xlsx` filename."""
    if not filename or not filename.strip():
        raise InventoryImportValidationError(
            CODE_UNSUPPORTED_FILE_TYPE,
            "A file with a name is required.",
            status_code=400,
        )
    extension = get_inventory_import_file_extension(filename)
    if extension not in ALLOWED_INVENTORY_IMPORT_EXTENSIONS:
        raise InventoryImportValidationError(
            CODE_UNSUPPORTED_FILE_TYPE,
            "Only .xlsx files exported from QuickBooks are supported.",
            status_code=400,
        )


def validate_inventory_import_content_type(content_type: str | None) -> None:
    """Lenient content-type gate.

    The filename extension is the real authority (browsers and test
    clients vary wildly on the MIME they attach to a `.xlsx`). We allow
    `None`, the empty string, generic binary types and the known xlsx
    types; only a *specific, clearly-wrong* type is rejected.
    """
    if content_type is None:
        return
    normalized = content_type.split(";", 1)[0].strip().lower()
    if normalized in INVENTORY_IMPORT_ALLOWED_CONTENT_TYPES:
        return
    raise InventoryImportValidationError(
        CODE_UNSUPPORTED_FILE_TYPE,
        f"Unsupported content type '{content_type}'. Upload an .xlsx file.",
        status_code=400,
    )


def validate_inventory_import_size(size_bytes: int) -> None:
    """Reject empty uploads and uploads above the configured ceiling."""
    if size_bytes <= 0:
        raise InventoryImportValidationError(
            CODE_EMPTY_FILE,
            "The uploaded file is empty.",
            status_code=400,
        )
    if size_bytes > MAX_INVENTORY_IMPORT_BYTES:
        raise InventoryImportValidationError(
            CODE_FILE_TOO_LARGE,
            (
                f"File is too large ({size_bytes} bytes); the limit is "
                f"{MAX_INVENTORY_IMPORT_BYTES} bytes."
            ),
            status_code=413,
        )


def validate_inventory_import_upload_metadata(
    filename: str,
    content_type: str | None,
    size_bytes: int,
) -> None:
    """Run every upload-level gate in order: name, type, size."""
    validate_inventory_import_filename(filename)
    validate_inventory_import_content_type(content_type)
    validate_inventory_import_size(size_bytes)


# --------------------------------------------------------------------- #
# Parser model (SUBFASE B)
# --------------------------------------------------------------------- #


@dataclass(slots=True)
class ParsedInventoryImportRow:
    """One QuickBooks item row, normalized.

    `parsed_quantity` is None when the quantity cell could not be parsed
    into a non-negative integer; `quantity_error` then carries the
    machine code (`INVALID_QUANTITY` / `QUANTITY_OUT_OF_RANGE`). The
    informational fields (`upc` … `unorderable`) are collected for
    completeness and possible future use; the preview surface only
    exposes a subset.
    """

    row_number: int
    raw_sku: str | None
    normalized_sku: str
    item_name: str | None
    parsed_quantity: int | None
    quantity_error: str | None = None
    upc: str | None = None
    department_name: str | None = None
    regular_price: float | None = None
    average_unit_cost: float | None = None
    unorderable: str | None = None


@dataclass(slots=True)
class ParsedInventoryImportWorkbook:
    """Result of parsing an uploaded workbook."""

    sheet_name: str
    header_row_number: int
    rows: list[ParsedInventoryImportRow] = field(default_factory=list)


# Normalized header label -> parsed-row field.
_HEADER_FIELD_MAP: dict[str, str] = {
    "item number": "sku",
    "item name": "item_name",
    "qty 1": "quantity",
    "upc": "upc",
    "department name": "department_name",
    "regular price": "regular_price",
    "average unit cost": "average_unit_cost",
    "unorderable": "unorderable",
}

_REQUIRED_HEADERS = ("item number", "qty 1")


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().lower().split())


def _coerce_optional_str(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _coerce_sku(value: object) -> str:
    """Normalize the Item Number cell to a trimmed string.

    QuickBooks stores item numbers as numbers, so `18163` arrives as an
    int (or sometimes a float like `18163.0`). Collapse integral floats
    to their integer string so the SKU matches what was stored as text.
    """
    if value is None:
        return ""
    if isinstance(value, bool):  # guard: bool is a subclass of int
        return str(value).strip()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _coerce_quantity(value: object) -> tuple[int | None, str | None]:
    """Parse a quantity cell into a non-negative int.

    Returns `(quantity, None)` on success or `(None, code)` on failure,
    where `code` is one of `INVALID_QUANTITY` / `QUANTITY_OUT_OF_RANGE`.
    """
    if value is None or (isinstance(value, str) and not value.strip()):
        return None, CODE_INVALID_QUANTITY

    if isinstance(value, bool):
        return None, CODE_INVALID_QUANTITY

    if isinstance(value, int):
        quantity = value
    elif isinstance(value, float):
        if not value.is_integer():
            return None, CODE_INVALID_QUANTITY
        quantity = int(value)
    else:
        text = str(value).strip()
        try:
            quantity = int(text)
        except ValueError:
            try:
                as_float = float(text)
            except ValueError:
                return None, CODE_INVALID_QUANTITY
            if not as_float.is_integer():
                return None, CODE_INVALID_QUANTITY
            quantity = int(as_float)

    if quantity < 0:
        return None, CODE_INVALID_QUANTITY
    if quantity > MAX_INVENTORY_IMPORT_QUANTITY:
        return None, CODE_QUANTITY_OUT_OF_RANGE
    return quantity, None


def _coerce_number(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _row_is_empty(values: tuple) -> bool:
    for cell in values:
        if cell is None:
            continue
        if isinstance(cell, str) and not cell.strip():
            continue
        return False
    return True


def _find_header(rows: list[tuple]) -> tuple[int, dict[str, int]] | None:
    """Locate the header row inside one sheet.

    Returns `(zero_based_index, {field: column_index})` for the first
    row that contains an `Item Number` column, or None if this sheet has
    no such header. The header is detected dynamically — never assumed
    to be row 1 or row 2.
    """
    for index, values in enumerate(rows):
        normalized = {_normalize_header(cell): col for col, cell in enumerate(values)}
        if "item number" not in normalized:
            continue
        column_map: dict[str, int] = {}
        for label, col in normalized.items():
            mapped = _HEADER_FIELD_MAP.get(label)
            if mapped is not None and mapped not in column_map:
                column_map[mapped] = col
        return index, column_map
    return None


def parse_quickbooks_inventory_workbook(
    data: bytes,
) -> ParsedInventoryImportWorkbook:
    """Parse a QuickBooks POS `.xlsx` export into normalized rows.

    DB-free and side-effect-free. Selects the first sheet that contains
    a header row (an `Item Number` column), ignores empty sheets, finds
    the header dynamically, maps columns by normalized name, skips
    entirely-empty data rows, and normalizes Item Number → SKU and
    Qty 1 → quantity.

    Raises `InventoryImportValidationError` for file-level problems:
      - unreadable / non-xlsx bytes  → UNSUPPORTED_FILE_TYPE
      - no sheet with a header row    → BAD_HEADERS
      - header present but Qty 1 col missing → BAD_HEADERS

    Per-row problems (empty SKU, bad/negative quantity) are NOT raised;
    they are recorded on the row (`normalized_sku == ""`,
    `quantity_error`) for the preview layer to surface.
    """
    if not data:
        raise InventoryImportValidationError(
            CODE_EMPTY_FILE, "The uploaded file is empty.", status_code=400
        )

    try:
        workbook = load_workbook(
            io.BytesIO(data), read_only=True, data_only=True
        )
    except Exception as exc:  # openpyxl raises a variety of errors
        raise InventoryImportValidationError(
            CODE_UNSUPPORTED_FILE_TYPE,
            "The file could not be read as an .xlsx workbook.",
            status_code=400,
        ) from exc

    try:
        for worksheet in workbook.worksheets:
            sheet_rows = [tuple(r) for r in worksheet.iter_rows(values_only=True)]
            if all(_row_is_empty(r) for r in sheet_rows):
                continue  # ignore empty sheet
            header = _find_header(sheet_rows)
            if header is None:
                continue  # has data but no Item Number header; keep looking
            header_index, column_map = header
            if "quantity" not in column_map:
                raise InventoryImportValidationError(
                    CODE_BAD_HEADERS,
                    "The 'Qty 1' column is required but was not found.",
                )
            parsed_rows = _parse_data_rows(
                sheet_rows, header_index, column_map
            )
            return ParsedInventoryImportWorkbook(
                sheet_name=worksheet.title,
                header_row_number=header_index + 1,
                rows=parsed_rows,
            )
    finally:
        workbook.close()

    raise InventoryImportValidationError(
        CODE_BAD_HEADERS,
        "No worksheet with an 'Item Number' header row was found.",
    )


def _parse_data_rows(
    sheet_rows: list[tuple],
    header_index: int,
    column_map: dict[str, int],
) -> list[ParsedInventoryImportRow]:
    def cell(values: tuple, key: str) -> object:
        col = column_map.get(key)
        if col is None or col >= len(values):
            return None
        return values[col]

    rows: list[ParsedInventoryImportRow] = []
    for offset, values in enumerate(sheet_rows[header_index + 1 :], start=1):
        if _row_is_empty(values):
            continue
        spreadsheet_row_number = header_index + 1 + offset

        raw_sku_cell = cell(values, "sku")
        normalized_sku = _coerce_sku(raw_sku_cell)
        parsed_quantity, quantity_error = _coerce_quantity(
            cell(values, "quantity")
        )

        rows.append(
            ParsedInventoryImportRow(
                row_number=spreadsheet_row_number,
                raw_sku=(
                    None if raw_sku_cell is None else str(raw_sku_cell)
                ),
                normalized_sku=normalized_sku,
                item_name=_coerce_optional_str(cell(values, "item_name")),
                parsed_quantity=parsed_quantity,
                quantity_error=quantity_error,
                upc=_coerce_optional_str(cell(values, "upc")),
                department_name=_coerce_optional_str(
                    cell(values, "department_name")
                ),
                regular_price=_coerce_number(cell(values, "regular_price")),
                average_unit_cost=_coerce_number(
                    cell(values, "average_unit_cost")
                ),
                unorderable=_coerce_optional_str(cell(values, "unorderable")),
            )
        )
    return rows


# --------------------------------------------------------------------- #
# Preview / diff (SUBFASE D) — analysis shared with confirm
# --------------------------------------------------------------------- #


@dataclass(slots=True)
class _AnalyzedRow:
    parsed: ParsedInventoryImportRow
    matched_variant_id: UUID | None
    matched_product_name: str | None
    current_on_hand: int | None
    quantity_reserved: int | None
    new_on_hand: int | None
    delta: int | None
    action: str
    errors: list[InventoryImportRowIssue]
    warnings: list[InventoryImportRowIssue]


def _duplicate_skus(workbook: ParsedInventoryImportWorkbook) -> set[str]:
    seen: dict[str, int] = {}
    for row in workbook.rows:
        if row.normalized_sku:
            seen[row.normalized_sku] = seen.get(row.normalized_sku, 0) + 1
    return {sku for sku, count in seen.items() if count > 1}


def _variants_by_sku(
    db: Session, skus: set[str]
) -> dict[str, ProductVariant]:
    if not skus:
        return {}
    stmt = select(ProductVariant).where(ProductVariant.sku.in_(skus))
    return {variant.sku: variant for variant in db.scalars(stmt).all()}


def _existing_item(
    db: Session, store_id: UUID, variant_id: UUID
) -> InventoryItem | None:
    stmt = select(InventoryItem).where(
        InventoryItem.store_id == store_id,
        InventoryItem.variant_id == variant_id,
    )
    return db.scalar(stmt)


def _analyze_workbook(
    db: Session,
    store_id: UUID,
    workbook: ParsedInventoryImportWorkbook,
) -> list[_AnalyzedRow]:
    """Diff every parsed row against the DB without mutating anything.

    Shared by preview (rendering) and confirm (the blocking-error gate).
    Reads `ProductVariant` (global SKU) and `InventoryItem`
    (store-scoped) only.
    """
    duplicates = _duplicate_skus(workbook)
    variants = _variants_by_sku(
        db, {r.normalized_sku for r in workbook.rows if r.normalized_sku}
    )

    analyzed: list[_AnalyzedRow] = []
    for row in workbook.rows:
        errors: list[InventoryImportRowIssue] = []
        warnings: list[InventoryImportRowIssue] = []
        matched_variant_id: UUID | None = None
        matched_product_name: str | None = None
        current_on_hand: int | None = None
        quantity_reserved: int | None = None
        new_on_hand: int | None = None
        delta: int | None = None
        action = ACTION_SKIP

        if not row.normalized_sku:
            errors.append(
                InventoryImportRowIssue(
                    code=CODE_MISSING_SKU,
                    message="Item Number (SKU) is empty.",
                )
            )
        if row.quantity_error == CODE_INVALID_QUANTITY:
            errors.append(
                InventoryImportRowIssue(
                    code=CODE_INVALID_QUANTITY,
                    message="Qty 1 is not a valid non-negative integer.",
                )
            )
        elif row.quantity_error == CODE_QUANTITY_OUT_OF_RANGE:
            errors.append(
                InventoryImportRowIssue(
                    code=CODE_QUANTITY_OUT_OF_RANGE,
                    message=(
                        "Qty 1 exceeds the maximum allowed "
                        f"({MAX_INVENTORY_IMPORT_QUANTITY})."
                    ),
                )
            )
        if row.normalized_sku and row.normalized_sku in duplicates:
            errors.append(
                InventoryImportRowIssue(
                    code=CODE_DUPLICATE_SKU_IN_FILE,
                    message="This SKU appears more than once in the file.",
                )
            )

        variant = (
            variants.get(row.normalized_sku) if row.normalized_sku else None
        )
        if row.normalized_sku and variant is None:
            errors.append(
                InventoryImportRowIssue(
                    code=CODE_VARIANT_NOT_FOUND,
                    message=(
                        "No product variant exists with this SKU; a "
                        "variant must be created before importing stock."
                    ),
                )
            )

        if variant is not None:
            matched_variant_id = variant.id
            matched_product_name = (
                variant.product.name if variant.product is not None else None
            )
            item = _existing_item(db, store_id, variant.id)
            if item is not None:
                current_on_hand = item.quantity_on_hand
                quantity_reserved = item.quantity_reserved
                if row.parsed_quantity is not None:
                    new_on_hand = row.parsed_quantity
                    delta = new_on_hand - current_on_hand
                    if item.quantity_reserved > new_on_hand:
                        errors.append(
                            InventoryImportRowIssue(
                                code=CODE_RESERVED_EXCEEDS_NEW_ONHAND,
                                message=(
                                    "New on-hand "
                                    f"({new_on_hand}) is below the "
                                    f"{item.quantity_reserved} units already "
                                    "reserved."
                                ),
                            )
                        )
                action = ACTION_UPDATE
            else:
                current_on_hand = 0
                quantity_reserved = 0
                if row.parsed_quantity is not None:
                    new_on_hand = row.parsed_quantity
                    delta = new_on_hand
                action = ACTION_CREATE_INVENTORY_ITEM

        if errors:
            action = ACTION_SKIP

        analyzed.append(
            _AnalyzedRow(
                parsed=row,
                matched_variant_id=matched_variant_id,
                matched_product_name=matched_product_name,
                current_on_hand=current_on_hand,
                quantity_reserved=quantity_reserved,
                new_on_hand=new_on_hand,
                delta=delta,
                action=action,
                errors=errors,
                warnings=warnings,
            )
        )
    return analyzed


def _to_preview_row(analyzed: _AnalyzedRow) -> InventoryImportPreviewRow:
    row = analyzed.parsed
    return InventoryImportPreviewRow(
        row_number=row.row_number,
        raw_sku=row.raw_sku,
        normalized_sku=row.normalized_sku,
        item_name=row.item_name,
        parsed_quantity=row.parsed_quantity,
        matched_variant_id=analyzed.matched_variant_id,
        matched_product_name=analyzed.matched_product_name,
        current_on_hand=analyzed.current_on_hand,
        quantity_reserved=analyzed.quantity_reserved,
        new_on_hand=analyzed.new_on_hand,
        delta=analyzed.delta,
        action=analyzed.action,
        errors=analyzed.errors,
        warnings=analyzed.warnings,
    )


def _summarize(analyzed: list[_AnalyzedRow]) -> InventoryImportSummary:
    return InventoryImportSummary(
        total_rows=len(analyzed),
        valid_rows=sum(1 for a in analyzed if not a.errors),
        rows_with_errors=sum(1 for a in analyzed if a.errors),
        rows_with_warnings=sum(1 for a in analyzed if a.warnings),
        to_update=sum(1 for a in analyzed if a.action == ACTION_UPDATE),
        to_create_inventory_item=sum(
            1 for a in analyzed if a.action == ACTION_CREATE_INVENTORY_ITEM
        ),
        to_skip=sum(1 for a in analyzed if a.action == ACTION_SKIP),
        blocking_error_count=sum(len(a.errors) for a in analyzed),
    )


def build_inventory_import_preview(
    db: Session,
    store_id: UUID,
    workbook: ParsedInventoryImportWorkbook,
) -> InventoryImportPreviewResponse:
    """Build the preview response. Performs zero DB writes."""
    analyzed = _analyze_workbook(db, store_id, workbook)
    return InventoryImportPreviewResponse(
        store_id=store_id,
        summary=_summarize(analyzed),
        rows=[_to_preview_row(a) for a in analyzed],
    )


# --------------------------------------------------------------------- #
# Confirm (SUBFASE E) — single all-or-nothing transaction
# --------------------------------------------------------------------- #


def confirm_inventory_import(
    db: Session,
    store_id: UUID,
    workbook: ParsedInventoryImportWorkbook,
    *,
    actor_user_id: UUID | None,
) -> InventoryImportConfirmResponse:
    """Apply the import inside one transaction.

    Re-validates server-side (never trusts a prior preview): if ANY row
    carries a blocking error the whole import is refused before any
    write. Then, for each applicable row, locks/creates the
    `InventoryItem`, sets `quantity_on_hand` to the parsed quantity, and
    appends an `InventoryLog` (movement_type=adjustment,
    reason="inventory_import") only when the delta is non-zero. A single
    commit lands everything; any failure rolls the entire import back.
    """
    analyzed = _analyze_workbook(db, store_id, workbook)
    if any(a.errors for a in analyzed):
        raise InventoryImportValidationError(
            CODE_BLOCKING_ERRORS,
            "The import has blocking errors and cannot be confirmed.",
        )

    updated_count = 0
    created_count = 0
    skipped_count = 0
    unchanged_count = 0
    log_count = 0

    # All-or-nothing: any failure inside the loop rolls back every prior
    # mutation in this transaction. The `InventoryImportValidationError`
    # branches re-raise as controlled errors; any other exception is a
    # genuine fault that must also leave the DB untouched.
    try:
        for analyzed_row in analyzed:
            if analyzed_row.action == ACTION_SKIP:
                skipped_count += 1
                continue

            variant_id = analyzed_row.matched_variant_id
            new_on_hand = analyzed_row.new_on_hand
            # Guarded by the blocking-error gate above; both are
            # populated for update/create actions.
            if variant_id is None or new_on_hand is None:
                skipped_count += 1
                continue

            item = _lock_item_for_store(db, store_id, variant_id)
            if item is not None:
                if item.quantity_reserved > new_on_hand:
                    # Re-checked under lock: a concurrent reservation
                    # could have pushed reserved past the new on-hand
                    # since the analysis read. Abort the whole import.
                    raise InventoryImportValidationError(
                        CODE_RESERVED_EXCEEDS_NEW_ONHAND,
                        (
                            "A row's new on-hand is below the units "
                            "already reserved; the import was rolled back."
                        ),
                    )
                delta = new_on_hand - item.quantity_on_hand
                item.quantity_on_hand = new_on_hand
                if delta != 0:
                    inventory_svc._write_inventory_log(
                        db,
                        item=item,
                        movement_type=InventoryMovementType.adjustment,
                        quantity_delta=delta,
                        quantity_after=item.quantity_on_hand,
                        actor_user_id=actor_user_id,
                        reason=INVENTORY_IMPORT_REASON,
                        reference_type=None,
                        reference_id=None,
                    )
                    log_count += 1
                    updated_count += 1
                else:
                    unchanged_count += 1
            else:
                item = InventoryItem(
                    store_id=store_id,
                    variant_id=variant_id,
                    quantity_on_hand=new_on_hand,
                    quantity_reserved=0,
                    reorder_threshold=0,
                    status=InventoryStatus.available,
                )
                db.add(item)
                db.flush()  # assign id so the log can reference it
                if new_on_hand != 0:
                    inventory_svc._write_inventory_log(
                        db,
                        item=item,
                        movement_type=InventoryMovementType.adjustment,
                        quantity_delta=new_on_hand,
                        quantity_after=new_on_hand,
                        actor_user_id=actor_user_id,
                        reason=INVENTORY_IMPORT_REASON,
                        reference_type=None,
                        reference_id=None,
                    )
                    log_count += 1
                created_count += 1

        inventory_svc._commit_or_translate(db)
    except Exception:
        db.rollback()
        raise

    return InventoryImportConfirmResponse(
        store_id=store_id,
        updated_count=updated_count,
        created_inventory_item_count=created_count,
        skipped_count=skipped_count,
        unchanged_count=unchanged_count,
        inventory_log_count=log_count,
    )


def _lock_item_for_store(
    db: Session, store_id: UUID, variant_id: UUID
) -> InventoryItem | None:
    """Lock the (store, variant) inventory row FOR UPDATE if it exists.

    Returns None when no row exists yet (the caller then creates one).
    Mirrors the locking contract used across `app.services.inventory`.
    """
    stmt = (
        select(InventoryItem)
        .where(
            InventoryItem.store_id == store_id,
            InventoryItem.variant_id == variant_id,
        )
        .with_for_update()
    )
    return db.scalar(stmt)
