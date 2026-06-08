"""Tests for the Excel (QuickBooks POS) inventory import — F2.27.8.

Three layers, mirroring the service structure:

  1. Upload plumbing + pure parser (no DB): extension / content-type /
     size gates, and `parse_quickbooks_inventory_workbook` against
     synthetic workbooks and the real sample export.
  2. Preview endpoint (read-only diff): RBAC / tenancy, duplicate and
     variant-not-found detection, on-hand/new/delta math, and the hard
     guarantee that preview writes nothing.
  3. Confirm endpoint (transactional apply): RBAC / tenancy, update vs.
     create, audit-log shape, delta-zero / reserved-exceeds handling,
     server-side re-validation and all-or-nothing rollback.

The real sample lives at `<repo>/excel/QB POS Inventory Items Export.xlsx`.
"""

import uuid
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Callable

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import func
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.db.models import InventoryItem
from app.db.models import InventoryLog
from app.db.models import InventoryMovementType
from app.db.models import InventoryStatus
from app.db.models import Product
from app.db.models import ProductVariant
from app.db.models import Store
from app.db.models import User
from app.db.models import UserRole
from app.services import inventory_import as imp
from tests.helpers.auth import auth_headers_for as _auth
from tests.helpers.auth import make_user as central_make_user


XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
SAMPLE_PATH = (
    Path(__file__).resolve().parents[2]
    / "excel"
    / "QB POS Inventory Items Export.xlsx"
)
DEFAULT_HEADERS = [
    "Item Number",
    "Item Name",
    "Qty 1",
    "UPC",
    "Department Name",
    "Regular Price",
    "Average Unit Cost",
    "Unorderable",
]


# --------------------------------------------------------------------- #
# Workbook builders / fixtures
# --------------------------------------------------------------------- #


def make_xlsx_bytes(
    data_rows: list[tuple],
    *,
    headers: list[str] | None = None,
    header_row: int = 2,
    extra_empty_sheets: bool = True,
) -> bytes:
    """Build an in-memory `.xlsx`.

    Header row defaults to row 2 (a blank row 1) to mirror the real
    QuickBooks export. Each data row is a tuple aligned to `headers`.
    """
    headers = headers or DEFAULT_HEADERS
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for col, label in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=label)
    for i, row in enumerate(data_rows):
        for col, value in enumerate(row, start=1):
            ws.cell(row=header_row + 1 + i, column=col, value=value)
    if extra_empty_sheets:
        wb.create_sheet("Sheet2")
        wb.create_sheet("Sheet3")
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _row(sku, qty, name="Item", dept="DEPT", upc="", price=9.99, cost=4.0,
         unorderable="No") -> tuple:
    return (sku, name, qty, upc, dept, price, cost, unorderable)


@pytest.fixture
def make_store(db_session: Session) -> Callable[..., Store]:
    def _create(code: str | None = None) -> Store:
        store = Store(name="Imp", code=code or f"imp-{uuid.uuid4().hex[:8]}")
        db_session.add(store)
        db_session.commit()
        db_session.refresh(store)
        return store

    return _create


@pytest.fixture
def make_user(db_session: Session, make_store) -> Callable[..., User]:
    def _create(role: UserRole, store_id: uuid.UUID | None = None) -> User:
        sid = None if role == UserRole.admin else (
            store_id if store_id is not None else make_store().id
        )
        return central_make_user(
            db_session, role=role, store_id=sid, full_name=f"Imp {role.value}"
        )

    return _create


@pytest.fixture
def make_variant(db_session: Session) -> Callable[..., ProductVariant]:
    def _create(sku: str, product_name: str = "Prod") -> ProductVariant:
        product = Product(name=f"{product_name}-{uuid.uuid4().hex[:6]}", category="vape")
        db_session.add(product)
        db_session.commit()
        db_session.refresh(product)
        variant = ProductVariant(
            product_id=product.id, sku=sku, price=Decimal("9.99")
        )
        db_session.add(variant)
        db_session.commit()
        db_session.refresh(variant)
        return variant

    return _create


@pytest.fixture
def make_item(db_session: Session) -> Callable[..., InventoryItem]:
    def _create(
        store: Store,
        variant: ProductVariant,
        quantity_on_hand: int = 5,
        quantity_reserved: int = 0,
        status: InventoryStatus = InventoryStatus.available,
    ) -> InventoryItem:
        item = InventoryItem(
            store_id=store.id,
            variant_id=variant.id,
            quantity_on_hand=quantity_on_hand,
            quantity_reserved=quantity_reserved,
            status=status,
        )
        db_session.add(item)
        db_session.commit()
        db_session.refresh(item)
        return item

    return _create


def _preview(client, store_id, data, *, headers, filename="inv.xlsx",
             content_type=XLSX_CONTENT_TYPE):
    return client.post(
        f"/stores/{store_id}/inventory/import/preview",
        files={"file": (filename, data, content_type)},
        headers=headers,
    )


def _confirm(client, store_id, data, *, headers, filename="inv.xlsx",
             content_type=XLSX_CONTENT_TYPE):
    return client.post(
        f"/stores/{store_id}/inventory/import/confirm",
        files={"file": (filename, data, content_type)},
        headers=headers,
    )


# ===================================================================== #
# SUBFASE A — upload plumbing
# ===================================================================== #


@pytest.mark.parametrize("name", ["a.xlsx", "A.XLSX", "data.XlSx"])
def test_filename_accepts_xlsx_case_insensitive(name):
    imp.validate_inventory_import_filename(name)  # no raise


@pytest.mark.parametrize("name", ["a.csv", "a.xls", "a.txt", "noext", ""])
def test_filename_rejects_non_xlsx(name):
    with pytest.raises(imp.InventoryImportValidationError) as exc:
        imp.validate_inventory_import_filename(name)
    assert exc.value.code in {
        imp.CODE_UNSUPPORTED_FILE_TYPE,
    }


def test_size_rejects_zero_and_over_limit():
    with pytest.raises(imp.InventoryImportValidationError) as zero:
        imp.validate_inventory_import_size(0)
    assert zero.value.code == imp.CODE_EMPTY_FILE

    with pytest.raises(imp.InventoryImportValidationError) as big:
        imp.validate_inventory_import_size(imp.MAX_INVENTORY_IMPORT_BYTES + 1)
    assert big.value.code == imp.CODE_FILE_TOO_LARGE


def test_size_accepts_within_limit():
    imp.validate_inventory_import_size(1)
    imp.validate_inventory_import_size(imp.MAX_INVENTORY_IMPORT_BYTES)


def test_content_type_accepts_xlsx_and_generic_and_none():
    imp.validate_inventory_import_content_type(XLSX_CONTENT_TYPE)
    imp.validate_inventory_import_content_type("application/octet-stream")
    imp.validate_inventory_import_content_type(None)
    imp.validate_inventory_import_content_type("")


def test_content_type_rejects_clearly_wrong():
    with pytest.raises(imp.InventoryImportValidationError) as exc:
        imp.validate_inventory_import_content_type("text/csv")
    assert exc.value.code == imp.CODE_UNSUPPORTED_FILE_TYPE


# ===================================================================== #
# SUBFASE B — pure parser
# ===================================================================== #


def test_parse_real_sample_headers_on_row_two():
    data = SAMPLE_PATH.read_bytes()
    wb = imp.parse_quickbooks_inventory_workbook(data)
    assert wb.sheet_name == "Sheet1"
    assert wb.header_row_number == 2
    assert len(wb.rows) == 4
    skus = {r.normalized_sku for r in wb.rows}
    assert skus == {"18163", "23118", "23803", "14629"}
    by_sku = {r.normalized_sku: r for r in wb.rows}
    assert by_sku["23803"].parsed_quantity == 11
    assert by_sku["23803"].department_name == "KRATOM"


def test_parse_normalizes_numeric_item_number_to_string():
    data = make_xlsx_bytes([_row(18163, 1)])
    wb = imp.parse_quickbooks_inventory_workbook(data)
    assert wb.rows[0].normalized_sku == "18163"
    assert isinstance(wb.rows[0].normalized_sku, str)


def test_parse_ignores_empty_sheets_and_finds_header_dynamically():
    data = make_xlsx_bytes([_row("SKU1", 3)], header_row=5)
    wb = imp.parse_quickbooks_inventory_workbook(data)
    assert wb.header_row_number == 5
    assert wb.rows[0].normalized_sku == "SKU1"


def test_parse_skips_fully_empty_rows():
    data = make_xlsx_bytes(
        [_row("SKU1", 1), (None, None, None, None, None, None, None, None),
         _row("SKU2", 2)]
    )
    wb = imp.parse_quickbooks_inventory_workbook(data)
    assert [r.normalized_sku for r in wb.rows] == ["SKU1", "SKU2"]


def test_parse_records_empty_sku_and_bad_quantity_as_row_state():
    data = make_xlsx_bytes([_row("", 1), _row("SKU2", "abc"),
                            _row("SKU3", -4)])
    wb = imp.parse_quickbooks_inventory_workbook(data)
    assert wb.rows[0].normalized_sku == ""
    assert wb.rows[1].parsed_quantity is None
    assert wb.rows[1].quantity_error == imp.CODE_INVALID_QUANTITY
    assert wb.rows[2].parsed_quantity is None
    assert wb.rows[2].quantity_error == imp.CODE_INVALID_QUANTITY


def test_parse_rejects_non_xlsx_bytes():
    with pytest.raises(imp.InventoryImportValidationError) as exc:
        imp.parse_quickbooks_inventory_workbook(b"not a zip file at all")
    assert exc.value.code == imp.CODE_UNSUPPORTED_FILE_TYPE


def test_parse_rejects_missing_item_number_header():
    data = make_xlsx_bytes(
        [("x", 1)], headers=["Wrong Header", "Qty 1"]
    )
    with pytest.raises(imp.InventoryImportValidationError) as exc:
        imp.parse_quickbooks_inventory_workbook(data)
    assert exc.value.code == imp.CODE_BAD_HEADERS


def test_parse_rejects_missing_qty_header():
    data = make_xlsx_bytes(
        [("x",)], headers=["Item Number", "Item Name"]
    )
    with pytest.raises(imp.InventoryImportValidationError) as exc:
        imp.parse_quickbooks_inventory_workbook(data)
    assert exc.value.code == imp.CODE_BAD_HEADERS


# ===================================================================== #
# SUBFASE D — preview endpoint
# ===================================================================== #


def test_preview_requires_file(client: TestClient, make_user):
    manager = make_user(UserRole.manager)
    resp = client.post(
        f"/stores/{manager.store_id}/inventory/import/preview",
        headers=_auth(manager),
    )
    assert resp.status_code == 422  # FastAPI: missing required UploadFile


def test_preview_rejects_unsupported_extension(client, make_user):
    manager = make_user(UserRole.manager)
    resp = _preview(
        client, manager.store_id, b"x", headers=_auth(manager),
        filename="data.csv", content_type="text/csv",
    )
    assert resp.status_code == 400
    assert resp.json()["detail"]["code"] == imp.CODE_UNSUPPORTED_FILE_TYPE


def test_preview_rejects_bad_headers(client, make_user):
    manager = make_user(UserRole.manager)
    data = make_xlsx_bytes([("x", 1)], headers=["Nope", "Qty 1"])
    resp = _preview(client, manager.store_id, data, headers=_auth(manager))
    assert resp.status_code == 422
    assert resp.json()["detail"]["code"] == imp.CODE_BAD_HEADERS


def test_preview_parses_real_sample(client, make_user):
    manager = make_user(UserRole.manager)
    data = SAMPLE_PATH.read_bytes()
    resp = _preview(client, manager.store_id, data, headers=_auth(manager))
    assert resp.status_code == 200
    body = resp.json()
    assert body["summary"]["total_rows"] == 4
    # No variants exist for these SKUs → all VARIANT_NOT_FOUND (blocking).
    assert body["summary"]["blocking_error_count"] == 4


def test_preview_detects_duplicate_sku_in_file(client, make_user, make_variant):
    manager = make_user(UserRole.manager)
    store_id = manager.store_id
    make_variant("DUP")
    data = make_xlsx_bytes([_row("DUP", 1), _row("DUP", 2)])
    resp = _preview(client, store_id, data, headers=_auth(manager))
    assert resp.status_code == 200
    codes = {
        issue["code"]
        for row in resp.json()["rows"]
        for issue in row["errors"]
    }
    assert imp.CODE_DUPLICATE_SKU_IN_FILE in codes


def test_preview_detects_variant_not_found(client, make_user):
    manager = make_user(UserRole.manager)
    data = make_xlsx_bytes([_row("MISSING-SKU", 3)])
    resp = _preview(client, manager.store_id, data, headers=_auth(manager))
    row = resp.json()["rows"][0]
    assert row["action"] == "skip"
    assert {i["code"] for i in row["errors"]} == {imp.CODE_VARIANT_NOT_FOUND}


def test_preview_detects_reserved_exceeds_new_onhand(
    client, db_session, make_user, make_variant, make_item
):
    manager = make_user(UserRole.manager)
    store = db_session.get(Store, manager.store_id)
    variant = make_variant("RSV")
    make_item(store, variant, quantity_on_hand=10, quantity_reserved=8)
    data = make_xlsx_bytes([_row("RSV", 3)])  # new on-hand 3 < reserved 8
    resp = _preview(client, store.id, data, headers=_auth(manager))
    row = resp.json()["rows"][0]
    assert {i["code"] for i in row["errors"]} == {
        imp.CODE_RESERVED_EXCEEDS_NEW_ONHAND
    }


def test_preview_computes_update_diff_and_action(
    client, db_session, make_user, make_variant, make_item
):
    manager = make_user(UserRole.manager)
    store = db_session.get(Store, manager.store_id)
    variant = make_variant("UPD")
    make_item(store, variant, quantity_on_hand=4)
    data = make_xlsx_bytes([_row("UPD", 10)])
    resp = _preview(client, store.id, data, headers=_auth(manager))
    row = resp.json()["rows"][0]
    assert row["action"] == "update"
    assert row["current_on_hand"] == 4
    assert row["new_on_hand"] == 10
    assert row["delta"] == 6
    assert row["matched_product_name"] is not None


def test_preview_create_action_when_no_item(
    client, make_user, make_variant
):
    manager = make_user(UserRole.manager)
    make_variant("NEW")
    data = make_xlsx_bytes([_row("NEW", 7)])
    resp = _preview(client, manager.store_id, data, headers=_auth(manager))
    row = resp.json()["rows"][0]
    assert row["action"] == "create_inventory_item"
    assert row["current_on_hand"] == 0
    assert row["new_on_hand"] == 7
    assert row["delta"] == 7


def test_preview_writes_nothing(
    client, db_session, make_user, make_variant
):
    manager = make_user(UserRole.manager)
    make_variant("NOWRITE")
    items_before = db_session.scalar(select(func.count()).select_from(InventoryItem))
    logs_before = db_session.scalar(select(func.count()).select_from(InventoryLog))
    data = make_xlsx_bytes([_row("NOWRITE", 9)])
    resp = _preview(client, manager.store_id, data, headers=_auth(manager))
    assert resp.status_code == 200
    assert db_session.scalar(select(func.count()).select_from(InventoryItem)) == items_before
    assert db_session.scalar(select(func.count()).select_from(InventoryLog)) == logs_before


def test_preview_blocks_staff(client, make_user):
    staff = make_user(UserRole.staff)
    data = make_xlsx_bytes([_row("S", 1)])
    resp = _preview(client, staff.store_id, data, headers=_auth(staff))
    assert resp.status_code == 403


def test_preview_blocks_cross_store(client, make_user, make_store):
    manager = make_user(UserRole.manager)
    other = make_store()
    data = make_xlsx_bytes([_row("S", 1)])
    resp = _preview(client, other.id, data, headers=_auth(manager))
    assert resp.status_code == 403


def test_preview_admin_any_store(client, make_user, make_store):
    admin = make_user(UserRole.admin)
    store = make_store()
    data = make_xlsx_bytes([_row("S", 1)])
    resp = _preview(client, store.id, data, headers=_auth(admin))
    assert resp.status_code == 200


def test_preview_owner_allowed_own_store(client, make_user):
    owner = make_user(UserRole.owner)
    data = make_xlsx_bytes([_row("S", 1)])
    resp = _preview(client, owner.store_id, data, headers=_auth(owner))
    assert resp.status_code == 200


def test_preview_anon_unauthorized(client, make_user):
    manager = make_user(UserRole.manager)
    data = make_xlsx_bytes([_row("S", 1)])
    resp = _preview(client, manager.store_id, data, headers={})
    assert resp.status_code == 401


# ===================================================================== #
# SUBFASE E — confirm endpoint
# ===================================================================== #


def test_confirm_blocks_staff(client, make_user):
    staff = make_user(UserRole.staff)
    data = make_xlsx_bytes([_row("S", 1)])
    resp = _confirm(client, staff.store_id, data, headers=_auth(staff))
    assert resp.status_code == 403


def test_confirm_blocks_cross_store(client, make_user, make_store):
    manager = make_user(UserRole.manager)
    other = make_store()
    data = make_xlsx_bytes([_row("S", 1)])
    resp = _confirm(client, other.id, data, headers=_auth(manager))
    assert resp.status_code == 403


def test_confirm_updates_existing_item_and_writes_log(
    client, db_session, make_user, make_variant, make_item
):
    manager = make_user(UserRole.manager)
    store = db_session.get(Store, manager.store_id)
    variant = make_variant("C-UPD")
    item = make_item(store, variant, quantity_on_hand=4)
    data = make_xlsx_bytes([_row("C-UPD", 10)])
    resp = _confirm(client, store.id, data, headers=_auth(manager))
    assert resp.status_code == 200
    body = resp.json()
    assert body["updated_count"] == 1
    assert body["inventory_log_count"] == 1

    db_session.expire_all()
    refreshed = db_session.get(InventoryItem, item.id)
    assert refreshed.quantity_on_hand == 10
    log = db_session.scalar(
        select(InventoryLog).where(InventoryLog.inventory_item_id == item.id)
    )
    assert log.movement_type == InventoryMovementType.adjustment
    assert log.quantity_delta == 6
    assert log.quantity_after == 10
    assert log.reason == "inventory_import"
    assert log.reference_type is None
    assert log.reference_id is None
    assert log.performed_by_user_id == manager.id


def test_confirm_creates_item_when_variant_exists_without_item(
    client, db_session, make_user, make_variant
):
    manager = make_user(UserRole.manager)
    store_id = manager.store_id
    variant = make_variant("C-NEW")
    data = make_xlsx_bytes([_row("C-NEW", 7)])
    resp = _confirm(client, store_id, data, headers=_auth(manager))
    assert resp.status_code == 200
    assert resp.json()["created_inventory_item_count"] == 1

    db_session.expire_all()
    item = db_session.scalar(
        select(InventoryItem).where(
            InventoryItem.store_id == store_id,
            InventoryItem.variant_id == variant.id,
        )
    )
    assert item is not None
    assert item.quantity_on_hand == 7
    assert item.quantity_reserved == 0
    assert item.status == InventoryStatus.available


def test_confirm_no_log_for_zero_delta(
    client, db_session, make_user, make_variant, make_item
):
    manager = make_user(UserRole.manager)
    store = db_session.get(Store, manager.store_id)
    variant = make_variant("C-SAME")
    make_item(store, variant, quantity_on_hand=5)
    data = make_xlsx_bytes([_row("C-SAME", 5)])  # delta 0
    resp = _confirm(client, store.id, data, headers=_auth(manager))
    body = resp.json()
    assert body["unchanged_count"] == 1
    assert body["updated_count"] == 0
    assert body["inventory_log_count"] == 0


def test_confirm_does_not_create_product_or_variant(
    client, db_session, make_user, make_variant
):
    manager = make_user(UserRole.manager)
    products_before = db_session.scalar(select(func.count()).select_from(Product))
    variants_before = db_session.scalar(select(func.count()).select_from(ProductVariant))
    make_variant("C-NOAUTO")  # one variant for the row to match
    data = make_xlsx_bytes([_row("C-NOAUTO", 3)])
    resp = _confirm(client, manager.store_id, data, headers=_auth(manager))
    assert resp.status_code == 200
    # Exactly the one variant (+ its product) we created above; the import
    # itself created none.
    assert db_session.scalar(select(func.count()).select_from(Product)) == products_before + 1
    assert db_session.scalar(select(func.count()).select_from(ProductVariant)) == variants_before + 1


def test_confirm_blocked_by_variant_not_found(client, make_user):
    manager = make_user(UserRole.manager)
    data = make_xlsx_bytes([_row("GHOST", 3)])
    resp = _confirm(client, manager.store_id, data, headers=_auth(manager))
    assert resp.status_code == 422
    assert resp.json()["detail"]["code"] == imp.CODE_BLOCKING_ERRORS


def test_confirm_blocked_by_reserved_exceeds(
    client, db_session, make_user, make_variant, make_item
):
    manager = make_user(UserRole.manager)
    store = db_session.get(Store, manager.store_id)
    variant = make_variant("C-RSV")
    make_item(store, variant, quantity_on_hand=10, quantity_reserved=8)
    data = make_xlsx_bytes([_row("C-RSV", 3)])
    resp = _confirm(client, store.id, data, headers=_auth(manager))
    assert resp.status_code == 422
    assert resp.json()["detail"]["code"] == imp.CODE_BLOCKING_ERRORS


def test_confirm_admin_any_store(client, make_user, make_store, make_variant):
    admin = make_user(UserRole.admin)
    store = make_store()
    make_variant("C-ADM")
    data = make_xlsx_bytes([_row("C-ADM", 4)])
    resp = _confirm(client, store.id, data, headers=_auth(admin))
    assert resp.status_code == 200
    assert resp.json()["created_inventory_item_count"] == 1


def test_confirm_revalidates_even_after_valid_preview(
    client, db_session, make_user, make_variant
):
    """A preview that passed must not let a now-invalid file confirm.

    Preview against an existing variant succeeds; we then delete the
    variant and confirm the SAME bytes — confirm must re-validate and
    refuse (VARIANT_NOT_FOUND → blocking).
    """
    manager = make_user(UserRole.manager)
    store_id = manager.store_id
    variant = make_variant("C-REVAL")
    data = make_xlsx_bytes([_row("C-REVAL", 4)])

    ok = _preview(client, store_id, data, headers=_auth(manager))
    assert ok.json()["summary"]["blocking_error_count"] == 0

    db_session.delete(variant)
    db_session.commit()

    resp = _confirm(client, store_id, data, headers=_auth(manager))
    assert resp.status_code == 422
    assert resp.json()["detail"]["code"] == imp.CODE_BLOCKING_ERRORS


def test_confirm_all_or_nothing_rollback(
    monkeypatch, db_session, make_user, make_variant, make_item
):
    """If a row fails mid-apply, NO row is persisted.

    We force the second log write to raise and assert the first row's
    update did not survive. Exercised at the service layer so we can
    inject the fault and inspect the rolled-back session directly.
    """
    manager = make_user(UserRole.manager)
    store = db_session.get(Store, manager.store_id)
    v1 = make_variant("RB-1")
    v2 = make_variant("RB-2")
    i1 = make_item(store, v1, quantity_on_hand=1)
    i2 = make_item(store, v2, quantity_on_hand=1)

    data = make_xlsx_bytes([_row("RB-1", 9), _row("RB-2", 9)])
    workbook = imp.parse_quickbooks_inventory_workbook(data)

    real_write = imp.inventory_svc._write_inventory_log
    calls = {"n": 0}

    def _boom(*args, **kwargs):
        calls["n"] += 1
        if calls["n"] == 2:
            raise RuntimeError("induced failure on second row")
        return real_write(*args, **kwargs)

    monkeypatch.setattr(imp.inventory_svc, "_write_inventory_log", _boom)

    with pytest.raises(RuntimeError):
        imp.confirm_inventory_import(
            db_session, store.id, workbook, actor_user_id=manager.id
        )

    db_session.expire_all()
    assert db_session.get(InventoryItem, i1.id).quantity_on_hand == 1
    assert db_session.get(InventoryItem, i2.id).quantity_on_hand == 1
    assert db_session.scalar(select(func.count()).select_from(InventoryLog)) == 0
